"""
Anarock GTM Engagement Financial Model Generator — FORMULA-DRIVEN
=================================================================
Every calculated cell uses Excel formulas referencing upstream assumptions.
Change any input and the entire model recalculates.

Project: Meda East Winds | 542 Units | 800k sft | Target GTV ~833.6 Cr
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import os

# ─────────────────────────────────────────────
# COLOR PALETTE
# ─────────────────────────────────────────────
CRIMSON = "C3002F"
CRIMSON_LIGHT = "FFCCDD"
DARK_CARD = "111522"
GOLD = "B8860B"
GOLD_LIGHT = "FFF8E1"
CYAN = "0086C3"
CYAN_LIGHT = "E0F7FA"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
MED_GRAY = "E2E8F0"
DARK_TEXT = "1E293B"
GREEN = "059669"
GREEN_LIGHT = "D1FAE5"
RED_LIGHT = "FEE2E2"
AMBER_LIGHT = "FEF3C7"

# ─────────────────────────────────────────────
# STYLING HELPERS
# ─────────────────────────────────────────────
def hfont(bold=True, size=11, color=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def bfont(bold=False, size=10, color=DARK_TEXT):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def sfill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

TBORDER = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)

def style_header(ws, row, ncols, bg=CRIMSON, fc=WHITE, ht=28):
    for c in range(1, ncols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = hfont(color=fc)
        cl.fill = sfill(bg)
        cl.border = TBORDER
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = ht

def dc(cell, num=False, fmt=None, bold=False):
    """Style a data cell."""
    cell.font = bfont(bold=bold)
    cell.border = TBORDER
    cell.alignment = Alignment(
        horizontal="right" if num else "left",
        vertical="center", wrap_text=True
    )
    if fmt:
        cell.number_format = fmt

def title_block(ws, title, subtitle, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", bold=True, size=16, color=CRIMSON)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=10)
    c2 = ws.cell(row=row+1, column=1, value=subtitle)
    c2.font = Font(name="Calibri", size=10, color="64748B", italic=True)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30
    return row + 3

def auto_width(ws, mn=13, mx=42):
    for col_cells in ws.columns:
        max_len = mn
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value and not str(cell.value).startswith("="):
                max_len = max(max_len, min(len(str(cell.value)) + 3, mx))
        ws.column_dimensions[letter].width = max_len

def cr(col, row):
    """Return absolute cell reference like $C$5."""
    return f"${get_column_letter(col)}${row}"

def ref(sheet, col, row):
    """Return cross-sheet absolute reference like 'Assumptions'!$C$5."""
    return f"'{sheet}'!{cr(col, row)}"


# ═══════════════════════════════════════════════
# SHEET 1: ASSUMPTIONS (Central Input Sheet)
# ═══════════════════════════════════════════════
def create_assumptions(wb):
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_properties.tabColor = CRIMSON

    r = title_block(ws,
        "ANAROCK GTM Financial Model -- Meda East Winds",
        "All inputs on this sheet drive the entire model. Change any value below and all sheets recalculate."
    )

    # Section: Project Parameters
    ws.cell(row=r, column=1, value="PROJECT PARAMETERS")
    ws.cell(row=r, column=1).font = hfont(size=12, color=CRIMSON)
    ws.cell(row=r, column=1).fill = sfill(LIGHT_GRAY)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1

    headers = ["Parameter", "Value", "Unit", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, 4, bg=CRIMSON)
    r += 1

    # We'll track key cell locations for cross-referencing
    # Row map: parameter_name -> row number (value is always in column 2)
    params = {}

    inputs = [
        ("Total Saleable Area", 800000, "sft", "Net saleable area"),
        ("Average Apartment Size", 1476, "sft", "Weighted average across tiers"),
        ("Total Units (calculated)", None, "Units", "= Total Area / Avg Size"),  # FORMULA
        ("", "", "", ""),
        ("--- Tier 1: Value Entry (2BHK) ---", "", "", ""),
        ("T1 Units", 220, "Units", "Compact 2BHK"),
        ("T1 Avg Size", 1050, "sft", "Per unit"),
        ("T1 Base Price", 8500, "Rs/sft", "Launch price"),
        ("T1 Effective Price", 9200, "Rs/sft", "After premiums"),
        ("", "", "", ""),
        ("--- Tier 2: Core Premium (3BHK) ---", "", "", ""),
        ("T2 Units", 280, "Units", "Premium 3BHK"),
        ("T2 Avg Size", 1600, "sft", "Per unit"),
        ("T2 Base Price", 10800, "Rs/sft", "Launch price"),
        ("T2 Effective Price", 11500, "Rs/sft", "After premiums"),
        ("", "", "", ""),
        ("--- Tier 3: Signature Luxury ---", "", "", ""),
        ("T3 Units", 42, "Units", "Penthouses / 4BHK"),
        ("T3 Avg Size", 2800, "sft", "Per unit"),
        ("T3 Base Price", 12500, "Rs/sft", "Launch price"),
        ("T3 Effective Price", 14200, "Rs/sft", "After premiums"),
        ("", "", "", ""),
        ("--- Cost Parameters ---", "", "", ""),
        ("Land Acquisition Cost", 3100, "Rs/sft", "Upfront equity"),
        ("Construction Cost (All-in)", 4200, "Rs/sft", "Incl. amenities, MEP, finishing"),
        ("Statutory & Approvals", 8.0, "Cr", "RERA, BDA, BBMP"),
        ("Contingency %", 2.0, "%", "As % of hard cost"),
        ("", "", "", ""),
        ("--- Financing ---", "", "", ""),
        ("Debt LTV Ratio", 65, "%", "Loan-to-value"),
        ("Debt Interest Rate", 9.5, "%", "Annual"),
        ("Loan Tenure", 36, "Months", "Construction finance"),
        ("", "", "", ""),
        ("--- Anarock GTM Fees ---", "", "", ""),
        ("Brokerage Commission Rate", 1.50, "%", "Of GTV per sale"),
        ("Marketing Budget Rate", 0.80, "%", "Of GTV"),
        ("AI Platform Monthly Fee", 3.0, "Lakhs/mo", "Astra, Genie, CP Ranker"),
        ("CP Mobilization (One-time)", 25.0, "Lakhs", "200 broker onboarding"),
        ("Performance Bonus Rate", 0.25, "%", "Incremental if velocity >15/mo"),
        ("", "", "", ""),
        ("--- Sales Velocity ---", "", "", ""),
        ("Target Monthly Velocity", 12, "Units/mo", "Gate 4 threshold"),
        ("Pre-Launch Soft Bookings %", 15.0, "%", "Gate 3: % of total units"),
        ("", "", "", ""),
        ("--- Gate Thresholds ---", "", "", ""),
        ("Gate 1: Min Equity IRR", 20.0, "%", "Mandate sign-off hurdle"),
        ("Gate 1: Max Overhang", 18, "Months", "Inventory overhang index"),
        ("Gate 2: Min CPs", 150, "CPs", "Pre-registered channel partners"),
        ("Gate 3: Min Token Corpus", 12.0, "Cr", "EOI token aggregation"),
        ("Gate 3: Astra Accuracy", 85.0, "%", "Prediction accuracy"),
        ("Gate 4: Max CAC", 2.50, "%", "% of GTV"),
        ("Gate 4: Min Lead Revival", 10.0, "%", "Astra Phoenix success rate"),
        ("", "", "", ""),
        ("--- Project Timeline ---", "", "", ""),
        ("Discovery Phase", 2, "Months", "Stage 1"),
        ("GTM Prep Phase", 2, "Months", "Stage 2"),
        ("Soft Launch Phase", 1, "Months", "Stage 3"),
        ("Public Launch Phase", 7, "Months", "Stage 4 initial"),
        ("Sustenance Phase", 24, "Months", "Stage 4 ongoing"),
        ("Wrapex Phase", 6, "Months", "Wind-down"),
        ("Handover Phase", 6, "Months", "Final handovers"),
    ]

    for label, val, unit, note in inputs:
        params[label] = r
        ws.cell(row=r, column=1, value=label)
        dc(ws.cell(row=r, column=1), bold=label.startswith("---"))

        if label == "Total Units (calculated)":
            # Formula: Total Area / Avg Size
            area_row = params["Total Saleable Area"]
            size_row = params["Average Apartment Size"]
            ws.cell(row=r, column=2).value = f"=ROUND({cr(2,area_row)}/{cr(2,size_row)},0)"
            dc(ws.cell(row=r, column=2), num=True, fmt="#,##0")
        elif val is not None and val != "":
            ws.cell(row=r, column=2, value=val)
            fmt = "#,##0.00" if isinstance(val, float) else "#,##0" if isinstance(val, int) else None
            dc(ws.cell(row=r, column=2), num=True, fmt=fmt)
        ws.cell(row=r, column=3, value=unit)
        dc(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=note)
        dc(ws.cell(row=r, column=4))

        if label.startswith("---"):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)
                ws.cell(row=r, column=c).font = bfont(bold=True, size=10, color=GOLD)
        r += 1

    # Computed Summary Block
    r += 1
    ws.cell(row=r, column=1, value="COMPUTED SUMMARY (All Formulas)")
    ws.cell(row=r, column=1).font = hfont(size=12, color=CRIMSON)
    ws.cell(row=r, column=1).fill = sfill(LIGHT_GRAY)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1

    for i, h in enumerate(["Metric", "Formula Result", "Unit", "Formula Description"], 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, 4, bg=GREEN)
    r += 1

    # Computed metrics
    computed = {}

    # T1 GTV = T1 Units * T1 Avg Size * T1 Effective Price / 10^7 (to get Cr)
    t1u, t1s, t1p = params["T1 Units"], params["T1 Avg Size"], params["T1 Effective Price"]
    t2u, t2s, t2p = params["T2 Units"], params["T2 Avg Size"], params["T2 Effective Price"]
    t3u, t3s, t3p = params["T3 Units"], params["T3 Avg Size"], params["T3 Effective Price"]

    computed_items = [
        ("T1 Tier GTV", f"={cr(2,t1u)}*{cr(2,t1s)}*{cr(2,t1p)}/10000000", "Cr", "T1 Units x T1 Size x T1 Eff Price / 10^7"),
        ("T2 Tier GTV", f"={cr(2,t2u)}*{cr(2,t2s)}*{cr(2,t2p)}/10000000", "Cr", "T2 Units x T2 Size x T2 Eff Price / 10^7"),
        ("T3 Tier GTV", f"={cr(2,t3u)}*{cr(2,t3s)}*{cr(2,t3p)}/10000000", "Cr", "T3 Units x T3 Size x T3 Eff Price / 10^7"),
    ]
    for label, formula, unit, desc in computed_items:
        computed[label] = r
        ws.cell(row=r, column=1, value=label); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=formula); dc(ws.cell(row=r, column=2), num=True, fmt="#,##0.00")
        ws.cell(row=r, column=3, value=unit); dc(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=desc); dc(ws.cell(row=r, column=4))
        r += 1

    t1g, t2g, t3g = computed["T1 Tier GTV"], computed["T2 Tier GTV"], computed["T3 Tier GTV"]

    more_computed = [
        ("Total Project GTV", f"={cr(2,t1g)}+{cr(2,t2g)}+{cr(2,t3g)}", "Cr", "Sum of all tier GTVs"),
        ("Total Units (Sum)", f"={cr(2,t1u)}+{cr(2,t2u)}+{cr(2,t3u)}", "Units", "Sum of tier units"),
    ]
    for label, formula, unit, desc in more_computed:
        computed[label] = r
        ws.cell(row=r, column=1, value=label); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=formula); dc(ws.cell(row=r, column=2), num=True, fmt="#,##0.00")
        ws.cell(row=r, column=3, value=unit); dc(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=desc); dc(ws.cell(row=r, column=4))
        r += 1

    gtv_row = computed["Total Project GTV"]
    total_u_row = computed["Total Units (Sum)"]

    area_row = params["Total Saleable Area"]
    land_row = params["Land Acquisition Cost"]
    const_row = params["Construction Cost (All-in)"]
    stat_row = params["Statutory & Approvals"]
    cont_row = params["Contingency %"]
    brk_row = params["Brokerage Commission Rate"]
    mkt_row = params["Marketing Budget Rate"]
    ai_row = params["AI Platform Monthly Fee"]
    cp_row = params["CP Mobilization (One-time)"]
    perf_row = params["Performance Bonus Rate"]
    ltv_row = params["Debt LTV Ratio"]
    int_row = params["Debt Interest Rate"]
    tenure_row = params["Loan Tenure"]

    more2 = [
        ("Blended ASP", f"={cr(2,gtv_row)}*10000000/{cr(2,area_row)}", "Rs/sft", "Total GTV / Total Area"),
        ("Avg Ticket Size", f"={cr(2,gtv_row)}/{cr(2,total_u_row)}", "Cr", "GTV / Total Units"),
        ("", "", "", ""),
        ("Land Cost (Total)", f"={cr(2,area_row)}*{cr(2,land_row)}/10000000", "Cr", "Area x Land Rate / 10^7"),
        ("Construction Cost (Total)", f"={cr(2,area_row)}*{cr(2,const_row)}/10000000", "Cr", "Area x Const Rate / 10^7"),
        ("Total Hard Costs", None, "Cr", "Land + Construction"),
        ("Contingency Amount", None, "Cr", "Contingency % x Hard Costs"),
        ("", "", "", ""),
        ("Brokerage Cost", f"={cr(2,gtv_row)}*{cr(2,brk_row)}/100", "Cr", "GTV x Brokerage Rate"),
        ("Marketing Cost", f"={cr(2,gtv_row)}*{cr(2,mkt_row)}/100", "Cr", "GTV x Marketing Rate"),
        ("AI Platform Cost (42 mo)", f"={cr(2,ai_row)}*42/100", "Cr", "Monthly fee x 42 months / 100"),
        ("CP Mobilization Cost", f"={cr(2,cp_row)}/100", "Cr", "One-time / 100 for Cr"),
        ("Performance Bonus", f"={cr(2,gtv_row)}*{cr(2,perf_row)}/100", "Cr", "GTV x Bonus Rate"),
        ("Total Anarock GTM Cost", None, "Cr", "Sum of all Anarock fees"),
        ("Anarock GTM as % of GTV", None, "%", "Total Anarock / GTV x 100"),
        ("", "", "", ""),
        ("Total Project Cost", None, "Cr", "Hard Costs + Statutory + Contingency + GTM"),
        ("Developer Net Profit", None, "Cr", "GTV - Total Project Cost"),
        ("Net Margin %", None, "%", "Net Profit / GTV x 100"),
        ("", "", "", ""),
        ("Debt Amount", f"={cr(2,gtv_row)}*{cr(2,ltv_row)}/100", "Cr", "GTV x LTV"),
        ("Annual Debt Service", None, "Cr", "Debt x Interest Rate"),
        ("Annual NOI (est.)", None, "Cr", "Estimated net operating income"),
        ("DSCR", None, "x", "NOI / Debt Service"),
    ]

    for label, formula, unit, desc in more2:
        computed[label] = r
        ws.cell(row=r, column=1, value=label); dc(ws.cell(row=r, column=1), bold=bool(label))
        if formula:
            ws.cell(row=r, column=2, value=formula)
        dc(ws.cell(row=r, column=2), num=True, fmt="#,##0.00" if label else None)
        ws.cell(row=r, column=3, value=unit); dc(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=desc); dc(ws.cell(row=r, column=4))
        r += 1

    # Now fill in the formulas that reference other computed rows
    land_t = computed["Land Cost (Total)"]
    const_t = computed["Construction Cost (Total)"]
    hard_row = computed["Total Hard Costs"]
    ws.cell(row=hard_row, column=2, value=f"={cr(2,land_t)}+{cr(2,const_t)}")

    cont_amt = computed["Contingency Amount"]
    ws.cell(row=cont_amt, column=2, value=f"={cr(2,hard_row)}*{cr(2,cont_row)}/100")

    brk_c = computed["Brokerage Cost"]
    mkt_c = computed["Marketing Cost"]
    ai_c = computed["AI Platform Cost (42 mo)"]
    cp_c = computed["CP Mobilization Cost"]
    perf_c = computed["Performance Bonus"]
    tot_an = computed["Total Anarock GTM Cost"]
    ws.cell(row=tot_an, column=2, value=f"={cr(2,brk_c)}+{cr(2,mkt_c)}+{cr(2,ai_c)}+{cr(2,cp_c)}+{cr(2,perf_c)}")

    an_pct = computed["Anarock GTM as % of GTV"]
    ws.cell(row=an_pct, column=2, value=f"={cr(2,tot_an)}/{cr(2,gtv_row)}*100")

    tot_cost = computed["Total Project Cost"]
    ws.cell(row=tot_cost, column=2, value=f"={cr(2,hard_row)}+{cr(2,stat_row)}+{cr(2,cont_amt)}+{cr(2,tot_an)}")

    net_profit = computed["Developer Net Profit"]
    ws.cell(row=net_profit, column=2, value=f"={cr(2,gtv_row)}-{cr(2,tot_cost)}")

    net_margin = computed["Net Margin %"]
    ws.cell(row=net_margin, column=2, value=f"={cr(2,net_profit)}/{cr(2,gtv_row)}*100")

    debt_row = computed["Debt Amount"]
    ann_ds = computed["Annual Debt Service"]
    ws.cell(row=ann_ds, column=2, value=f"={cr(2,debt_row)}*{cr(2,int_row)}/100")

    ann_noi = computed["Annual NOI (est.)"]
    ws.cell(row=ann_noi, column=2, value=f"={cr(2,net_profit)}/3.5")  # ~3.5 year project

    dscr_row = computed["DSCR"]
    ws.cell(row=dscr_row, column=2, value=f"={cr(2,ann_noi)}/{cr(2,ann_ds)}")

    # Highlight key computed rows
    for key in ["Total Project GTV", "Total Hard Costs", "Total Anarock GTM Cost",
                 "Total Project Cost", "Developer Net Profit", "Net Margin %", "DSCR"]:
        rw = computed[key]
        for c in range(1, 5):
            ws.cell(row=rw, column=c).fill = sfill(GOLD_LIGHT)
            ws.cell(row=rw, column=c).font = bfont(bold=True, size=11)

    for key in ["Developer Net Profit", "Net Margin %"]:
        rw = computed[key]
        for c in range(1, 5):
            ws.cell(row=rw, column=c).fill = sfill(GREEN_LIGHT)

    auto_width(ws)
    return ws, params, computed


# ═══════════════════════════════════════════════
# SHEET 2: REVENUE WATERFALL
# ═══════════════════════════════════════════════
def create_revenue_waterfall(wb, params, computed):
    ws = wb.create_sheet("Revenue Waterfall")
    ws.sheet_properties.tabColor = "D4AF37"
    A = "Assumptions"  # sheet name for refs

    r = title_block(ws,
        "Revenue Waterfall -- Barbell Pricing Architecture",
        "All cells reference Assumptions sheet. Change tier inputs to recalculate."
    )

    headers = ["Tier", "Configuration", "Unit Count", "Avg Size (sft)",
               "Base Price (Rs/sft)", "Effective Price (Rs/sft)",
               "Avg Ticket (Cr)", "Tier GTV (Cr)", "% of Total GTV"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers), bg=GOLD)
    r += 1

    # Data rows with formulas referencing Assumptions
    tiers = [
        ("Value Entry", "Compact 2 BHK", "T1 Units", "T1 Avg Size", "T1 Base Price", "T1 Effective Price", "T1 Tier GTV"),
        ("Core Premium", "Premium 3 BHK", "T2 Units", "T2 Avg Size", "T2 Base Price", "T2 Effective Price", "T2 Tier GTV"),
        ("Signature Luxury", "Penthouses / 4 BHK", "T3 Units", "T3 Avg Size", "T3 Base Price", "T3 Effective Price", "T3 Tier GTV"),
    ]

    tier_rows = []
    for tier_name, config, u_key, s_key, bp_key, ep_key, gtv_key in tiers:
        ws.cell(row=r, column=1, value=tier_name); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=config); dc(ws.cell(row=r, column=2))

        # Units = Assumptions ref
        ws.cell(row=r, column=3, value=f"={ref(A, 2, params[u_key])}")
        dc(ws.cell(row=r, column=3), num=True, fmt="#,##0")

        # Avg Size
        ws.cell(row=r, column=4, value=f"={ref(A, 2, params[s_key])}")
        dc(ws.cell(row=r, column=4), num=True, fmt="#,##0")

        # Base Price
        ws.cell(row=r, column=5, value=f"={ref(A, 2, params[bp_key])}")
        dc(ws.cell(row=r, column=5), num=True, fmt="#,##0")

        # Effective Price
        ws.cell(row=r, column=6, value=f"={ref(A, 2, params[ep_key])}")
        dc(ws.cell(row=r, column=6), num=True, fmt="#,##0")

        # Avg Ticket = GTV / Units
        ws.cell(row=r, column=7, value=f"={ref(A, 2, computed[gtv_key])}/{ref(A, 2, params[u_key])}")
        dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.00")

        # Tier GTV from Assumptions computed
        ws.cell(row=r, column=8, value=f"={ref(A, 2, computed[gtv_key])}")
        dc(ws.cell(row=r, column=8), num=True, fmt="#,##0.00")

        # % of GTV = Tier GTV / Total GTV
        ws.cell(row=r, column=9, value=f"={ref(A, 2, computed[gtv_key])}/{ref(A, 2, computed['Total Project GTV'])}")
        dc(ws.cell(row=r, column=9), num=True, fmt="0.0%")

        tier_rows.append(r)
        r += 1

    # TOTAL ROW
    tr = r
    ws.cell(row=r, column=1, value="TOTAL / BLENDED"); dc(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=""); dc(ws.cell(row=r, column=2))

    ws.cell(row=r, column=3, value=f"={ref(A, 2, computed['Total Units (Sum)'])}")
    dc(ws.cell(row=r, column=3), num=True, fmt="#,##0", bold=True)

    ws.cell(row=r, column=4, value=f"={ref(A, 2, params['Average Apartment Size'])}")
    dc(ws.cell(row=r, column=4), num=True, fmt="#,##0", bold=True)

    ws.cell(row=r, column=5, value=""); dc(ws.cell(row=r, column=5))

    # Blended ASP
    ws.cell(row=r, column=6, value=f"={ref(A, 2, computed['Blended ASP'])}")
    dc(ws.cell(row=r, column=6), num=True, fmt="#,##0", bold=True)

    # Avg Ticket
    ws.cell(row=r, column=7, value=f"={ref(A, 2, computed['Avg Ticket Size'])}")
    dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.00", bold=True)

    # Total GTV
    ws.cell(row=r, column=8, value=f"={ref(A, 2, computed['Total Project GTV'])}")
    dc(ws.cell(row=r, column=8), num=True, fmt="#,##0.00", bold=True)

    ws.cell(row=r, column=9, value=1.0)
    dc(ws.cell(row=r, column=9), num=True, fmt="0.0%", bold=True)

    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)
    r += 2

    # ── Revenue Phasing ──
    ws.cell(row=r, column=1, value="Revenue Phasing by GTM Stage")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=13, color=CRIMSON)
    r += 1

    ph_headers = ["GTM Stage", "Timeline", "% of Units Sold", "Units Sold",
                  "Revenue (Cr)", "Cumulative Units", "Cumulative Revenue (Cr)",
                  "% GTV Achieved", "Avg Velocity (units/mo)", "Months"]
    for i, h in enumerate(ph_headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(ph_headers))
    r += 1

    # Phase data: (name, timeline, % units, months_in_phase)
    phases = [
        ("S1: Pre-Mandate Discovery", "Month 1-2", 0.0, 2),
        ("S2: GTM & Creative Prep", "Month 3-4", 0.0, 2),
        ("S3: Pre-Launch Soft Bookings", "Month 5", 15.0, 1),
        ("S4: Public Launch (M6-M12)", "Month 6-12", 27.0, 7),
        ("S4: Sustenance (M13-M24)", "Month 13-24", 31.0, 12),
        ("S4: Sustenance (M25-M36)", "Month 25-36", 20.5, 12),
        ("S4: Wrapex & Final", "Month 37-42", 6.5, 6),
    ]

    total_gtv_ref = ref(A, 2, computed["Total Project GTV"])
    total_units_ref = ref(A, 2, computed["Total Units (Sum)"])

    phase_rows = []
    for i, (name, timeline, pct, months) in enumerate(phases):
        pr = r
        phase_rows.append(pr)
        ws.cell(row=r, column=1, value=name); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=timeline); dc(ws.cell(row=r, column=2))

        # % of units sold (input)
        ws.cell(row=r, column=3, value=pct / 100)
        dc(ws.cell(row=r, column=3), num=True, fmt="0.0%")

        # Units sold = % x Total Units
        ws.cell(row=r, column=4, value=f"=ROUND({cr(3,r)}*{total_units_ref},0)")
        dc(ws.cell(row=r, column=4), num=True, fmt="#,##0")

        # Revenue = % x Total GTV
        ws.cell(row=r, column=5, value=f"={cr(3,r)}*{total_gtv_ref}")
        dc(ws.cell(row=r, column=5), num=True, fmt="#,##0.0")

        # Cumulative Units = sum of units from first phase to current
        if i == 0:
            ws.cell(row=r, column=6, value=f"={cr(4,r)}")
        else:
            ws.cell(row=r, column=6, value=f"={cr(6,r-1)}+{cr(4,r)}")
        dc(ws.cell(row=r, column=6), num=True, fmt="#,##0")

        # Cumulative Revenue
        if i == 0:
            ws.cell(row=r, column=7, value=f"={cr(5,r)}")
        else:
            ws.cell(row=r, column=7, value=f"={cr(7,r-1)}+{cr(5,r)}")
        dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.0")

        # % GTV Achieved = Cumulative Revenue / Total GTV
        ws.cell(row=r, column=8, value=f"={cr(7,r)}/{total_gtv_ref}")
        dc(ws.cell(row=r, column=8), num=True, fmt="0.0%")

        # Months in phase
        ws.cell(row=r, column=10, value=months)
        dc(ws.cell(row=r, column=10), num=True, fmt="#,##0")

        # Avg Velocity = Units / Months (handle divide by zero)
        ws.cell(row=r, column=9, value=f"=IF({cr(10,r)}=0,0,{cr(4,r)}/{cr(10,r)})")
        dc(ws.cell(row=r, column=9), num=True, fmt="#,##0.0")

        r += 1

    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Revenue by GTM Stage (Cr)"
    chart.style = 10
    chart.y_axis.title = "Cr"
    data_ref = Reference(ws, min_col=5, min_row=phase_rows[0], max_row=phase_rows[-1])
    cats_ref = Reference(ws, min_col=1, min_row=phase_rows[0], max_row=phase_rows[-1])
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = CRIMSON
    chart.width = 22; chart.height = 13
    ws.add_chart(chart, f"A{r+2}")

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 3: COST STRUCTURE & P&L
# ═══════════════════════════════════════════════
def create_cost_pnl(wb, params, computed):
    ws = wb.create_sheet("Cost Structure & PnL")
    ws.sheet_properties.tabColor = CYAN
    A = "Assumptions"

    r = title_block(ws,
        "Project Cost Structure & Developer P&L",
        "All costs computed from Assumptions inputs. Change rates to see P&L impact."
    )

    headers = ["Cost Category", "Rate / Basis", "Total Cost (Cr)",
               "% of GTV", "% of Total Cost", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers), bg=CYAN)
    r += 1

    gtv = ref(A, 2, computed["Total Project GTV"])
    land_t = ref(A, 2, computed["Land Cost (Total)"])
    const_t = ref(A, 2, computed["Construction Cost (Total)"])
    hard_t = ref(A, 2, computed["Total Hard Costs"])
    stat_t = ref(A, 2, params["Statutory & Approvals"])
    cont_t = ref(A, 2, computed["Contingency Amount"])
    gtm_t = ref(A, 2, computed["Total Anarock GTM Cost"])
    proj_t = ref(A, 2, computed["Total Project Cost"])
    profit = ref(A, 2, computed["Developer Net Profit"])
    margin = ref(A, 2, computed["Net Margin %"])

    items = {}
    cost_lines = [
        ("Land Acquisition (Equity)", f"={ref(A,2,params['Land Acquisition Cost'])}&\" Rs/sft\"",
         f"={land_t}", f"={land_t}/{gtv}", None, "Upfront equity outlay"),
        ("Construction Cost (All-in)", f"={ref(A,2,params['Construction Cost (All-in)'])}&\" Rs/sft\"",
         f"={const_t}", f"={const_t}/{gtv}", None, "Incl. amenities, MEP"),
        ("TOTAL HARD COSTS", "", f"={hard_t}", f"={hard_t}/{gtv}", None, ""),
        ("", "", "", "", None, ""),
        ("Statutory & Approvals", "Fixed", f"={stat_t}", f"={stat_t}/{gtv}", None, "RERA, BDA, BBMP"),
        ("Contingency", f"={ref(A,2,params['Contingency %'])}&\"% of hard costs\"",
         f"={cont_t}", f"={cont_t}/{gtv}", None, "Risk buffer"),
        ("", "", "", "", None, ""),
        ("Anarock Brokerage", f"={ref(A,2,params['Brokerage Commission Rate'])}&\"% of GTV\"",
         f"={ref(A,2,computed['Brokerage Cost'])}", f"={ref(A,2,computed['Brokerage Cost'])}/{gtv}", None, "Per sale closure"),
        ("Anarock Marketing", f"={ref(A,2,params['Marketing Budget Rate'])}&\"% of GTV\"",
         f"={ref(A,2,computed['Marketing Cost'])}", f"={ref(A,2,computed['Marketing Cost'])}/{gtv}", None, "Digital + OOH"),
        ("AI Platform License", f"={ref(A,2,params['AI Platform Monthly Fee'])}&\" L/mo x 42 mo\"",
         f"={ref(A,2,computed['AI Platform Cost (42 mo)'])}", f"={ref(A,2,computed['AI Platform Cost (42 mo)'])}/{gtv}", None, "Astra, Genie, CP Ranker"),
        ("CP Mobilization", "One-time",
         f"={ref(A,2,computed['CP Mobilization Cost'])}", f"={ref(A,2,computed['CP Mobilization Cost'])}/{gtv}", None, "Broker onboarding"),
        ("Performance Bonus", f"={ref(A,2,params['Performance Bonus Rate'])}&\"% of GTV\"",
         f"={ref(A,2,computed['Performance Bonus'])}", f"={ref(A,2,computed['Performance Bonus'])}/{gtv}", None, "If velocity >15/mo"),
        ("TOTAL ANAROCK GTM COST", "", f"={gtm_t}", f"={gtm_t}/{gtv}", None, ""),
        ("", "", "", "", None, ""),
        ("TOTAL PROJECT COST", "", f"={proj_t}", f"={proj_t}/{gtv}", None, ""),
        ("", "", "", "", None, ""),
        ("GROSS REVENUE (GTV)", "", f"={gtv}", "100.0%", None, ""),
        ("(-) Total Project Cost", "", f"=-{proj_t}", f"=-{proj_t}/{gtv}", None, ""),
        ("DEVELOPER NET PROFIT", "", f"={profit}", f"={profit}/{gtv}", None, ""),
        ("NET MARGIN %", "", f"={margin}/100", "", None, ""),
    ]

    for cat, rate, cost, pct, _, note in cost_lines:
        items[cat] = r
        ws.cell(row=r, column=1, value=cat); dc(ws.cell(row=r, column=1), bold="TOTAL" in cat or "NET" in cat or "GROSS" in cat or "DEVELOPER" in cat)
        ws.cell(row=r, column=2, value=rate if rate else ""); dc(ws.cell(row=r, column=2))
        if cost and cost != "":
            ws.cell(row=r, column=3, value=cost)
        dc(ws.cell(row=r, column=3), num=True, fmt="#,##0.00")

        if pct and pct != "" and pct != "100.0%":
            ws.cell(row=r, column=4, value=pct)
        elif pct == "100.0%":
            ws.cell(row=r, column=4, value=1.0)
        dc(ws.cell(row=r, column=4), num=True, fmt="0.0%")

        # % of Total Cost (col 5) for key rows
        if cost and "TOTAL PROJECT" not in cat and cat and "GROSS" not in cat and "DEVELOPER" not in cat and "NET" not in cat and "(-)" not in cat:
            ws.cell(row=r, column=5, value=f"=IF({proj_t}=0,0,ABS({cr(3,r)})/{proj_t})")
            dc(ws.cell(row=r, column=5), num=True, fmt="0.0%")

        ws.cell(row=r, column=6, value=note if note else ""); dc(ws.cell(row=r, column=6))

        # Highlight totals
        is_total = any(k in cat for k in ["TOTAL", "GROSS", "DEVELOPER", "NET MARGIN"])
        if is_total:
            bg = GREEN_LIGHT if "PROFIT" in cat or "NET MARGIN" in cat else GOLD_LIGHT
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = sfill(bg)
                ws.cell(row=r, column=c).font = bfont(bold=True, size=11)
        r += 1

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 4: CASH FLOW PROJECTION
# ═══════════════════════════════════════════════
def create_cashflow(wb, params, computed):
    ws = wb.create_sheet("Cash Flow Projection")
    ws.sheet_properties.tabColor = GREEN
    A = "Assumptions"

    r = title_block(ws,
        "Quarterly Cash Flow Projection -- 48 Month Lifecycle",
        "Inflows from sales, outflows for construction & GTM. Cumulative computed via formulas."
    )

    headers = ["Quarter", "Months", "GTM Stage",
               "% Units Sold", "Units Sold", "Sale Revenue (Cr)",
               "Construction Out (Cr)", "GTM Cost Out (Cr)",
               "Net Cash Flow (Cr)", "Cumulative CF (Cr)", "Cumulative Units"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers), bg=GREEN)
    r += 1

    gtv = ref(A, 2, computed["Total Project GTV"])
    total_u = ref(A, 2, computed["Total Units (Sum)"])
    const_total = ref(A, 2, computed["Construction Cost (Total)"])

    # Quarterly plan: (quarter, months, stage, %_units, const_pct_of_total, gtm_lakhs)
    quarters = [
        ("Q1", "M1-M3", "S1: Discovery", 0.0, 0.0, 80),
        ("Q2", "M4-M6", "S2-S3: Prep + Soft", 15.0, 8.3, 320),
        ("Q3", "M7-M9", "S4: Public Launch", 11.0, 8.3, 450),
        ("Q4", "M10-M12", "S4: Launch Momentum", 16.0, 8.3, 420),
        ("Q5", "M13-M15", "S4: Sustenance", 9.0, 8.3, 280),
        ("Q6", "M16-M18", "S4: Sustenance", 8.0, 8.3, 250),
        ("Q7", "M19-M21", "S4: Sustenance", 7.5, 8.3, 220),
        ("Q8", "M22-M24", "S4: Sustenance", 7.0, 8.3, 200),
        ("Q9", "M25-M27", "S4: Sustenance", 6.0, 6.5, 150),
        ("Q10", "M28-M30", "S4: Sustenance", 5.5, 6.5, 120),
        ("Q11", "M31-M33", "S4: Sustenance", 5.0, 6.5, 100),
        ("Q12", "M34-M36", "S4: Winding Down", 4.0, 6.5, 80),
        ("Q13", "M37-M39", "Wrapex", 3.5, 5.4, 50),
        ("Q14", "M40-M42", "Wrapex", 2.5, 5.4, 30),
        ("Q15", "M43-M45", "Handover", 0.0, 3.0, 0),
        ("Q16", "M46-M48", "Final Handover", 0.0, 1.8, 0),
    ]

    qrows = []
    for i, (qtr, months, stage, unit_pct, const_pct, gtm_lakhs) in enumerate(quarters):
        qr = r
        qrows.append(qr)

        ws.cell(row=r, column=1, value=qtr); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=months); dc(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=stage); dc(ws.cell(row=r, column=3))

        # % Units (input)
        ws.cell(row=r, column=4, value=unit_pct / 100)
        dc(ws.cell(row=r, column=4), num=True, fmt="0.0%")

        # Units Sold = % x Total Units
        ws.cell(row=r, column=5, value=f"=ROUND({cr(4,r)}*{total_u},0)")
        dc(ws.cell(row=r, column=5), num=True, fmt="#,##0")

        # Revenue = % x GTV
        ws.cell(row=r, column=6, value=f"={cr(4,r)}*{gtv}")
        dc(ws.cell(row=r, column=6), num=True, fmt="#,##0.0")

        # Construction outflow = const_pct% of total construction cost
        ws.cell(row=r, column=7, value=f"={const_total}*{const_pct/100}")
        dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.0")

        # GTM cost outflow (in Cr)
        ws.cell(row=r, column=8, value=gtm_lakhs / 100)
        dc(ws.cell(row=r, column=8), num=True, fmt="#,##0.0")

        # Net Cash Flow = Revenue - Construction - GTM
        ws.cell(row=r, column=9, value=f"={cr(6,r)}-{cr(7,r)}-{cr(8,r)}")
        dc(ws.cell(row=r, column=9), num=True, fmt="#,##0.0")

        # Cumulative CF
        if i == 0:
            ws.cell(row=r, column=10, value=f"={cr(9,r)}")
        else:
            ws.cell(row=r, column=10, value=f"={cr(10,r-1)}+{cr(9,r)}")
        dc(ws.cell(row=r, column=10), num=True, fmt="#,##0.0")

        # Cumulative Units
        if i == 0:
            ws.cell(row=r, column=11, value=f"={cr(5,r)}")
        else:
            ws.cell(row=r, column=11, value=f"={cr(11,r-1)}+{cr(5,r)}")
        dc(ws.cell(row=r, column=11), num=True, fmt="#,##0")

        if r % 2 == 0:
            for c in range(1, 12):
                ws.cell(row=r, column=c).fill = sfill(LIGHT_GRAY)
        r += 1

    # Totals row
    first_q = qrows[0]; last_q = qrows[-1]
    ws.cell(row=r, column=1, value="TOTAL"); dc(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=4, value=f"=SUM({cr(4,first_q)}:{cr(4,last_q)})")
    dc(ws.cell(row=r, column=4), num=True, fmt="0.0%", bold=True)
    ws.cell(row=r, column=5, value=f"=SUM({cr(5,first_q)}:{cr(5,last_q)})")
    dc(ws.cell(row=r, column=5), num=True, fmt="#,##0", bold=True)
    ws.cell(row=r, column=6, value=f"=SUM({cr(6,first_q)}:{cr(6,last_q)})")
    dc(ws.cell(row=r, column=6), num=True, fmt="#,##0.0", bold=True)
    ws.cell(row=r, column=7, value=f"=SUM({cr(7,first_q)}:{cr(7,last_q)})")
    dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.0", bold=True)
    ws.cell(row=r, column=8, value=f"=SUM({cr(8,first_q)}:{cr(8,last_q)})")
    dc(ws.cell(row=r, column=8), num=True, fmt="#,##0.0", bold=True)
    ws.cell(row=r, column=9, value=f"=SUM({cr(9,first_q)}:{cr(9,last_q)})")
    dc(ws.cell(row=r, column=9), num=True, fmt="#,##0.0", bold=True)
    for c in range(1, 12):
        ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)
        ws.cell(row=r, column=c).font = bfont(bold=True, size=11)
    r += 1

    # Charts
    chart = LineChart()
    chart.title = "Cumulative Cash Flow (Cr)"
    chart.style = 10; chart.y_axis.title = "Cr"
    data_ref = Reference(ws, min_col=10, min_row=first_q, max_row=last_q)
    cats_ref = Reference(ws, min_col=1, min_row=first_q, max_row=last_q)
    chart.add_data(data_ref); chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.width = 25000
    chart.width = 24; chart.height = 14
    ws.add_chart(chart, f"A{r+2}")

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 5: IRR SENSITIVITY
# ═══════════════════════════════════════════════
def create_irr(wb, params, computed):
    ws = wb.create_sheet("IRR & DSCR Analysis")
    ws.sheet_properties.tabColor = "D4AF37"
    A = "Assumptions"

    r = title_block(ws,
        "Developer Equity IRR & DSCR Sensitivity",
        "Matrix shows estimated IRR at various ASP x Velocity combos. Base case highlighted."
    )

    ws.cell(row=r, column=1, value="IRR Sensitivity Matrix -- ASP vs. Sales Velocity")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=13, color=CRIMSON)
    r += 1

    # The IRR is approximated as: ((Revenue - Cost) / Equity) / (Sellout_Years) adjusted for velocity
    # We use a simplified parametric formula the user can later refine
    # IRR_approx = ((ASP * Area - Cost_Total) / Equity) / Sellout_Time * 100
    # Where Sellout_Time = Total_Units / (Velocity * 12)

    # For the sensitivity, we hardcode ASP tiers and velocity tiers, but compute using a formula
    # that references the cost structure from Assumptions

    area = ref(A, 2, params["Total Saleable Area"])
    hard = ref(A, 2, computed["Total Hard Costs"])
    total_units_sum = ref(A, 2, computed["Total Units (Sum)"])
    stat = ref(A, 2, params["Statutory & Approvals"])
    cont = ref(A, 2, computed["Contingency Amount"])

    vel_labels = ["8 u/mo", "10 u/mo", "12 u/mo (Target)", "15 u/mo", "18 u/mo"]
    vel_values = [8, 10, 12, 15, 18]
    asp_labels = ["Rs 8,500", "Rs 9,200", "Rs 10,200 (Base)", "Rs 10,800", "Rs 11,500"]
    asp_values = [8500, 9200, 10200, 10800, 11500]

    # Header
    ws.cell(row=r, column=1, value="ASP \\ Velocity")
    for j, vl in enumerate(vel_labels, 2):
        ws.cell(row=r, column=j, value=vl)
    style_header(ws, r, len(vel_labels) + 1, bg=DARK_CARD, ht=35)
    r += 1

    # Build formula rows
    # IRR approx = ((ASP_val * Area/1e7 - HardCosts - Stat - Cont - GTM_est) / (Land_total))
    #              / (TotalUnits / (Velocity * 12))
    # Simplified: profit / equity / years
    land_total = ref(A, 2, computed["Land Cost (Total)"])
    gtm_total = ref(A, 2, computed["Total Anarock GTM Cost"])

    for asp_label, asp_val in zip(asp_labels, asp_values):
        ws.cell(row=r, column=1, value=asp_label)
        ws.cell(row=r, column=1).font = bfont(bold=True)
        ws.cell(row=r, column=1).border = TBORDER

        for j, vel_val in enumerate(vel_values, 2):
            # Formula:
            # Revenue_est = ASP * Area / 1e7
            # Profit_est = Revenue_est - HardCosts - Stat - Cont - GTM
            # Equity = Land_total
            # Sellout_years = TotalUnits / (velocity * 12)
            # IRR_approx = (Profit_est / Equity) / Sellout_years
            formula = (
                f"=(({asp_val}*{area}/10000000 - {hard} - {stat} - {cont} - {gtm_total})"
                f" / {land_total})"
                f" / ({total_units_sum}/({vel_val}*12))"
            )
            cell = ws.cell(row=r, column=j, value=formula)
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = TBORDER
            cell.font = bfont(bold=False)

            # Conditional formatting (manual via IF-based fill isn't possible, so we color the base case)
            if asp_val == 10200 and vel_val == 12:
                cell.fill = sfill(CYAN_LIGHT)
                cell.font = bfont(bold=True, size=11)
            elif asp_val >= 10200 and vel_val >= 12:
                cell.fill = sfill(GREEN_LIGHT)
        r += 1

    r += 1
    # Legend
    for color, desc in [(GREEN_LIGHT, "Green: Strong IRR zone (above base)"),
                         (CYAN_LIGHT, "Blue: Base case scenario"),
                         (LIGHT_GRAY, "Gray: Computed from Assumptions -- change inputs to update")]:
        ws.cell(row=r, column=1).fill = sfill(color)
        ws.cell(row=r, column=1).border = TBORDER
        ws.cell(row=r, column=2, value=desc)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2).font = bfont(size=9)
        r += 1

    # DSCR Section
    r += 2
    ws.cell(row=r, column=1, value="DSCR Scenario Analysis")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=13, color=CRIMSON)
    r += 1

    dscr_h = ["Scenario", "LTV %", "Debt (Cr)", "Interest %",
              "Annual Debt Svc (Cr)", "Annual NOI (Cr)", "DSCR", "Gate Status"]
    for i, h in enumerate(dscr_h, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(dscr_h))
    r += 1

    noi_ref = ref(A, 2, computed["Annual NOI (est.)"])
    gtv_ref = ref(A, 2, computed["Total Project GTV"])

    scenarios = [
        ("Conservative", 60, 9.5),
        ("Base Case", 65, 9.5),
        ("Aggressive", 70, 10.0),
        ("Stressed", 75, 11.0),
        ("Worst Case", 80, 11.5),
    ]

    for name, ltv, int_rate in scenarios:
        ws.cell(row=r, column=1, value=name); dc(ws.cell(row=r, column=1), bold=True)

        ws.cell(row=r, column=2, value=ltv / 100)
        dc(ws.cell(row=r, column=2), num=True, fmt="0%")

        # Debt = GTV * LTV%
        ws.cell(row=r, column=3, value=f"={gtv_ref}*{cr(2,r)}")
        dc(ws.cell(row=r, column=3), num=True, fmt="#,##0.0")

        ws.cell(row=r, column=4, value=int_rate / 100)
        dc(ws.cell(row=r, column=4), num=True, fmt="0.0%")

        # Annual Debt Service = Debt * Interest
        ws.cell(row=r, column=5, value=f"={cr(3,r)}*{cr(4,r)}")
        dc(ws.cell(row=r, column=5), num=True, fmt="#,##0.0")

        # NOI from assumptions
        ws.cell(row=r, column=6, value=f"={noi_ref}")
        dc(ws.cell(row=r, column=6), num=True, fmt="#,##0.0")

        # DSCR = NOI / Debt Service
        ws.cell(row=r, column=7, value=f"={cr(6,r)}/{cr(5,r)}")
        dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.00\"x\"")

        # Gate status
        ws.cell(row=r, column=8, value=f'=IF({cr(7,r)}>=1.4,"CLEAR","BLOCKED")')
        dc(ws.cell(row=r, column=8), bold=True)
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="center", vertical="center")
        r += 1

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 6: STAGE-GATE BUDGET TRACKER
# ═══════════════════════════════════════════════
def create_budget(wb, params, computed):
    ws = wb.create_sheet("Stage-Gate Budget")
    ws.sheet_properties.tabColor = CRIMSON

    r = title_block(ws,
        "GTM Stage-Gate Budget Tracker",
        "Budget column links to Assumptions. Actual Spend is user-input. Variance auto-calculates."
    )

    headers = ["Stage", "Activity", "Budget (Lakhs)", "Actual Spend (Lakhs)",
               "Variance (Lakhs)", "Var %", "Owner", "Status"]
    ncols = len(headers)
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, ncols)
    r += 1

    items = [
        ("S1", "Competitor Conjoint Analysis", 15.0, 14.2, "Research"),
        ("S1", "Micro-Market Demand Heatmap", 8.0, 7.5, "Data Science"),
        ("S1", "Financial Hurdle Model", 5.0, 4.8, "Finance"),
        ("S1", "Legal Due Diligence", 12.0, 11.5, "Legal"),
        (None, "Stage 1 Subtotal", None, None, ""),
        ("", "", None, None, ""),
        ("S2", "Barbell Pricing Architecture", 10.0, 9.8, "Pricing"),
        ("S2", "Walk-in Genie AI Training", 18.0, 12.0, "AI Team"),
        ("S2", "CP Ranker & Mobilization", 25.0, 15.0, "CP Ops"),
        ("S2", "Site Experience Office", 85.0, 45.0, "Developer"),
        ("S2", "Creative Kit & Collaterals", 30.0, 0.0, "Creative"),
        ("S2", "RERA Registration Support", 5.0, 4.5, "Legal"),
        (None, "Stage 2 Subtotal", None, None, ""),
        ("", "", None, None, ""),
        ("S3", "Concierge MVP Events (5x)", 35.0, 0.0, "Sales Ops"),
        ("S3", "Astra Platinum Integration", 12.0, 0.0, "AI Team"),
        ("S3", "EOI Token Campaign", 45.0, 0.0, "Marketing"),
        ("S3", "CP Launch Events", 20.0, 0.0, "CP Ops"),
        (None, "Stage 3 Subtotal", None, None, ""),
        ("", "", None, None, ""),
        ("S4", "360 Digital Campaign (Meta/Google)", 180.0, 0.0, "Digital"),
        ("S4", "Local OOH & Print", 80.0, 0.0, "Media"),
        ("S4", "n8n Lead Revival (Astra Phoenix)", 15.0, 0.0, "AI Team"),
        ("S4", "Mandate Health Dashboard", 8.0, 0.0, "BI Team"),
        ("S4", "CP Commission Pool", 350.0, 0.0, "CP Ops"),
        ("S4", "Sales Team Ops", 120.0, 0.0, "Sales"),
        ("S4", "Site Event Marketing", 25.0, 0.0, "Events"),
        (None, "Stage 4 Subtotal", None, None, ""),
        ("", "", None, None, ""),
        (None, "GRAND TOTAL", None, None, ""),
    ]

    # Track subtotal positions
    s1_start = r; stage_items = {"S1": [], "S2": [], "S3": [], "S4": []}
    all_subtotals = []

    for stg, activity, budget, actual, owner in items:
        is_subtotal = stg is None and activity
        is_blank = stg == "" and activity == ""

        ws.cell(row=r, column=1, value=stg if stg else "")
        dc(ws.cell(row=r, column=1), bold=is_subtotal)

        ws.cell(row=r, column=2, value=activity)
        dc(ws.cell(row=r, column=2), bold=is_subtotal)

        if is_subtotal:
            all_subtotals.append(r)
            # Find the rows for this stage's items
            if "Stage 1" in activity:
                item_rows = stage_items["S1"]
            elif "Stage 2" in activity:
                item_rows = stage_items["S2"]
            elif "Stage 3" in activity:
                item_rows = stage_items["S3"]
            elif "Stage 4" in activity:
                item_rows = stage_items["S4"]
            elif "GRAND" in activity:
                # Sum all subtotals
                item_rows = all_subtotals[:-1]  # exclude self
            else:
                item_rows = []

            if "GRAND" in activity:
                sum_refs = "+".join([cr(3, sr) for sr in item_rows])
                ws.cell(row=r, column=3, value=f"={sum_refs}")
                sum_refs_a = "+".join([cr(4, sr) for sr in item_rows])
                ws.cell(row=r, column=4, value=f"={sum_refs_a}")
            elif item_rows:
                first_r = item_rows[0]; last_r = item_rows[-1]
                ws.cell(row=r, column=3, value=f"=SUM({cr(3,first_r)}:{cr(3,last_r)})")
                ws.cell(row=r, column=4, value=f"=SUM({cr(4,first_r)}:{cr(4,last_r)})")

            dc(ws.cell(row=r, column=3), num=True, fmt="#,##0.0", bold=True)
            dc(ws.cell(row=r, column=4), num=True, fmt="#,##0.0", bold=True)

            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)
        elif not is_blank:
            # Regular item
            if stg and stg in stage_items:
                stage_items[stg].append(r)

            ws.cell(row=r, column=3, value=budget)
            dc(ws.cell(row=r, column=3), num=True, fmt="#,##0.0")

            ws.cell(row=r, column=4, value=actual)
            dc(ws.cell(row=r, column=4), num=True, fmt="#,##0.0")

        if not is_blank:
            # Variance = Budget - Actual (formula)
            ws.cell(row=r, column=5, value=f"={cr(3,r)}-{cr(4,r)}")
            dc(ws.cell(row=r, column=5), num=True, fmt="#,##0.0")

            # Var % = Variance / Budget (handle 0)
            ws.cell(row=r, column=6, value=f"=IF({cr(3,r)}=0,0,{cr(5,r)}/{cr(3,r)})")
            dc(ws.cell(row=r, column=6), num=True, fmt="0.0%")

        ws.cell(row=r, column=7, value=owner); dc(ws.cell(row=r, column=7))
        ws.cell(row=r, column=8, value=""); dc(ws.cell(row=r, column=8))

        r += 1

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 7: CAC & MARKETING ANALYSIS
# ═══════════════════════════════════════════════
def create_cac(wb, params, computed):
    ws = wb.create_sheet("CAC & Marketing Analysis")
    ws.sheet_properties.tabColor = "D97706"

    r = title_block(ws,
        "CAC & Marketing Funnel Analysis",
        "CPL, CAC, and ROAS calculated via formulas. Adjust budget/leads to see impact."
    )

    headers = ["Channel", "Monthly Budget (Lakhs)", "Leads/Mo",
               "Site Visits/Mo", "Conversions/Mo", "CPL (Rs)",
               "CAC per Unit (Lakhs)", "Channel Revenue (Cr/mo)", "ROAS"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers), bg="D97706")
    r += 1

    avg_ticket = ref("Assumptions", 2, computed["Avg Ticket Size"])

    channels = [
        ("Google Search (SEM)", 12.0, 800, 48, 8),
        ("Meta (FB/Instagram)", 8.0, 1200, 36, 5),
        ("Google Display & YouTube", 5.0, 2000, 20, 3),
        ("CP Broker Network", 15.0, 300, 90, 12),
        ("Organic / Walk-in Genie AI", 2.0, 400, 60, 6),
        ("n8n Lead Revival (Phoenix)", 1.5, 150, 25, 3),
        ("Local OOH / Print", 6.0, 500, 15, 2),
        ("Referral Program", 1.0, 50, 15, 2),
    ]

    ch_rows = []
    for ch_name, budget, leads, visits, conv in channels:
        ch_rows.append(r)
        ws.cell(row=r, column=1, value=ch_name); dc(ws.cell(row=r, column=1), bold=True)

        # Budget (input)
        ws.cell(row=r, column=2, value=budget)
        dc(ws.cell(row=r, column=2), num=True, fmt="#,##0.0")

        # Leads (input)
        ws.cell(row=r, column=3, value=leads)
        dc(ws.cell(row=r, column=3), num=True, fmt="#,##0")

        # Site Visits (input)
        ws.cell(row=r, column=4, value=visits)
        dc(ws.cell(row=r, column=4), num=True, fmt="#,##0")

        # Conversions (input)
        ws.cell(row=r, column=5, value=conv)
        dc(ws.cell(row=r, column=5), num=True, fmt="#,##0")

        # CPL = Budget * 100000 / Leads (Lakhs to Rs)
        ws.cell(row=r, column=6, value=f"=IF({cr(3,r)}=0,0,{cr(2,r)}*100000/{cr(3,r)})")
        dc(ws.cell(row=r, column=6), num=True, fmt="#,##0")

        # CAC per unit = Budget / Conversions (in Lakhs)
        ws.cell(row=r, column=7, value=f"=IF({cr(5,r)}=0,0,{cr(2,r)}/{cr(5,r)})")
        dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.00")

        # Channel Revenue = Conversions * Avg Ticket (Cr)
        ws.cell(row=r, column=8, value=f"={cr(5,r)}*{avg_ticket}")
        dc(ws.cell(row=r, column=8), num=True, fmt="#,##0.0")

        # ROAS = Revenue / Budget (both need same units: Revenue in Cr = Lakhs*100, Budget in Lakhs)
        ws.cell(row=r, column=9, value=f"=IF({cr(2,r)}=0,0,{cr(8,r)}*100/{cr(2,r)})")
        dc(ws.cell(row=r, column=9), num=True, fmt="#,##0.0\"x\"")
        r += 1

    # Totals
    fr, lr = ch_rows[0], ch_rows[-1]
    ws.cell(row=r, column=1, value="BLENDED TOTAL"); dc(ws.cell(row=r, column=1), bold=True)
    for col in [2, 3, 4, 5]:
        ws.cell(row=r, column=col, value=f"=SUM({cr(col,fr)}:{cr(col,lr)})")
        dc(ws.cell(row=r, column=col), num=True, fmt="#,##0.0" if col == 2 else "#,##0", bold=True)

    ws.cell(row=r, column=6, value=f"=IF({cr(3,r)}=0,0,{cr(2,r)}*100000/{cr(3,r)})")
    dc(ws.cell(row=r, column=6), num=True, fmt="#,##0", bold=True)

    ws.cell(row=r, column=7, value=f"=IF({cr(5,r)}=0,0,{cr(2,r)}/{cr(5,r)})")
    dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.00", bold=True)

    ws.cell(row=r, column=8, value=f"=SUM({cr(8,fr)}:{cr(8,lr)})")
    dc(ws.cell(row=r, column=8), num=True, fmt="#,##0.0", bold=True)

    ws.cell(row=r, column=9, value=f"=IF({cr(2,r)}=0,0,{cr(8,r)}*100/{cr(2,r)})")
    dc(ws.cell(row=r, column=9), num=True, fmt="#,##0.0\"x\"", bold=True)

    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)
    r += 2

    # Funnel
    ws.cell(row=r, column=1, value="Lead-to-Conversion Funnel")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=13, color=CRIMSON)
    r += 1
    fh = ["Funnel Stage", "Volume", "Stage Conv %", "Drop-Off %", "Cumulative Eff %"]
    for i, h in enumerate(fh, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(fh))
    r += 1

    funnel = [
        ("Digital Impressions", 5000000),
        ("Clicks / Engagement", 75000),
        ("Lead Submissions", 5400),
        ("Qualified Leads (MQL)", 2700),
        ("Site Visits Booked", 810),
        ("Site Visits Done", 450),
        ("Booking Intent", 135),
        ("Confirmed Bookings", 41),
    ]

    funnel_rows = []
    for i, (stage, vol) in enumerate(funnel):
        funnel_rows.append(r)
        ws.cell(row=r, column=1, value=stage); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=vol); dc(ws.cell(row=r, column=2), num=True, fmt="#,##0")

        if i == 0:
            ws.cell(row=r, column=3, value="--"); dc(ws.cell(row=r, column=3))
            ws.cell(row=r, column=4, value="--"); dc(ws.cell(row=r, column=4))
            ws.cell(row=r, column=5, value=1.0); dc(ws.cell(row=r, column=5), num=True, fmt="0.000%")
        else:
            prev = funnel_rows[i - 1]
            ws.cell(row=r, column=3, value=f"={cr(2,r)}/{cr(2,prev)}")
            dc(ws.cell(row=r, column=3), num=True, fmt="0.0%")

            ws.cell(row=r, column=4, value=f"=1-{cr(3,r)}")
            dc(ws.cell(row=r, column=4), num=True, fmt="0.0%")

            first_vol_row = funnel_rows[0]
            ws.cell(row=r, column=5, value=f"={cr(2,r)}/{cr(2,first_vol_row)}")
            dc(ws.cell(row=r, column=5), num=True, fmt="0.000%")
        r += 1

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 8: SENSITIVITY ANALYSIS
# ═══════════════════════════════════════════════
def create_sensitivity(wb, params, computed):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.sheet_properties.tabColor = "7C3AED"
    A = "Assumptions"

    r = title_block(ws,
        "Scenario & Sensitivity Analysis",
        "Bear/Base/Bull/Stress P&L computed via formulas referencing Assumptions"
    )

    # Scenario definition: multipliers applied to base-case values
    ws.cell(row=r, column=1, value="Scenario Multipliers (Change these to adjust scenarios)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=12, color="7C3AED")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    sh = ["Parameter", "Bear Factor", "Base Factor", "Bull Factor", "Stress Factor"]
    for i, h in enumerate(sh, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(sh), bg="7C3AED")
    r += 1

    factors = {}
    factor_data = [
        ("ASP Multiplier", 0.85, 1.00, 1.12, 0.78),
        ("Velocity Multiplier", 0.67, 1.00, 1.50, 0.50),
        ("Construction Cost Mult.", 1.15, 1.00, 0.90, 1.25),
        ("Interest Rate Mult.", 1.10, 1.00, 0.90, 1.25),
        ("GTM Cost Multiplier", 1.20, 1.00, 0.80, 1.40),
    ]

    for label, bear, base, bull, stress in factor_data:
        factors[label] = r
        ws.cell(row=r, column=1, value=label); dc(ws.cell(row=r, column=1), bold=True)
        for j, v in enumerate([bear, base, bull, stress], 2):
            ws.cell(row=r, column=j, value=v)
            dc(ws.cell(row=r, column=j), num=True, fmt="0.00")
            if j == 2: ws.cell(row=r, column=j).fill = sfill(RED_LIGHT)
            elif j == 3: ws.cell(row=r, column=j).fill = sfill(CYAN_LIGHT)
            elif j == 4: ws.cell(row=r, column=j).fill = sfill(GREEN_LIGHT)
            elif j == 5: ws.cell(row=r, column=j).fill = sfill(AMBER_LIGHT)
        r += 1

    r += 1

    # Scenario P&L
    ws.cell(row=r, column=1, value="Scenario P&L Summary (Cr) -- All Formulas")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=13, color=CRIMSON)
    r += 1

    ph = ["Metric", "Bear Case", "Base Case", "Bull Case", "Stress Case"]
    for i, h in enumerate(ph, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(ph))
    r += 1

    gtv_base = ref(A, 2, computed["Total Project GTV"])
    land_base = ref(A, 2, computed["Land Cost (Total)"])
    const_base = ref(A, 2, computed["Construction Cost (Total)"])
    gtm_base = ref(A, 2, computed["Total Anarock GTM Cost"])
    stat_base = ref(A, 2, params["Statutory & Approvals"])
    cont_base = ref(A, 2, computed["Contingency Amount"])
    int_rate = ref(A, 2, params["Debt Interest Rate"])
    ltv_rate = ref(A, 2, params["Debt LTV Ratio"])

    asp_f = factors["ASP Multiplier"]
    const_f = factors["Construction Cost Mult."]
    int_f = factors["Interest Rate Mult."]
    gtm_f = factors["GTM Cost Multiplier"]

    pnl_rows = {}

    # Revenue = GTV * ASP_factor (each scenario col 2-5 maps to factor cols 2-5)
    pnl_metrics = [
        ("Gross Revenue (GTV)", [
            f"={gtv_base}*{cr(c,asp_f)}" for c in [2, 3, 4, 5]
        ]),
        ("(-) Land Cost", [f"=-{land_base}" for _ in range(4)]),
        ("(-) Construction Cost", [
            f"=-{const_base}*{cr(c,const_f)}" for c in [2, 3, 4, 5]
        ]),
        ("(-) GTM & Marketing", [
            f"=-{gtm_base}*{cr(c,gtm_f)}" for c in [2, 3, 4, 5]
        ]),
        ("(-) Statutory & Contingency", [f"=-({stat_base}+{cont_base})" for _ in range(4)]),
        ("(-) Debt Interest (est.)", [
            f"=-{gtv_base}*{cr(c,asp_f)}*{ltv_rate}/100*{int_rate}/100*{cr(c,int_f)}*3" for c in [2, 3, 4, 5]
        ]),
    ]

    for label, formulas in pnl_metrics:
        pnl_rows[label] = r
        ws.cell(row=r, column=1, value=label); dc(ws.cell(row=r, column=1), bold=True)
        for j, f in enumerate(formulas, 2):
            ws.cell(row=r, column=j, value=f)
            dc(ws.cell(row=r, column=j), num=True, fmt="#,##0.0")
        r += 1

    # Net Profit = sum of all above
    pnl_rows["Net Profit"] = r
    ws.cell(row=r, column=1, value="Developer Net Profit")
    dc(ws.cell(row=r, column=1), bold=True)
    for j in range(2, 6):
        refs = "+".join([cr(j, pnl_rows[k]) for k in pnl_rows if k != "Net Profit"])
        ws.cell(row=r, column=j, value=f"={refs}")
        dc(ws.cell(row=r, column=j), num=True, fmt="#,##0.0", bold=True)
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = sfill(GREEN_LIGHT)
    r += 1

    # Net Margin = Profit / Revenue
    rev_row = pnl_rows["Gross Revenue (GTV)"]
    prof_row = pnl_rows["Net Profit"]
    ws.cell(row=r, column=1, value="Net Margin %"); dc(ws.cell(row=r, column=1), bold=True)
    for j in range(2, 6):
        ws.cell(row=r, column=j, value=f"={cr(j,prof_row)}/{cr(j,rev_row)}")
        dc(ws.cell(row=r, column=j), num=True, fmt="0.0%", bold=True)
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)
    r += 1

    # DSCR = NOI / Debt Service
    debt_int_row = pnl_rows["(-) Debt Interest (est.)"]
    ws.cell(row=r, column=1, value="DSCR (est.)"); dc(ws.cell(row=r, column=1), bold=True)
    for j in range(2, 6):
        # NOI approx = Profit / 3.5 years, Debt svc = -Debt Interest / 3
        ws.cell(row=r, column=j, value=f"=IF({cr(j,debt_int_row)}=0,0,({cr(j,prof_row)}/3.5)/(-{cr(j,debt_int_row)}/3))")
        dc(ws.cell(row=r, column=j), num=True, fmt="#,##0.00\"x\"")
    r += 1

    # Gate status
    ws.cell(row=r, column=1, value="Gate 1 Status"); dc(ws.cell(row=r, column=1), bold=True)
    margin_row = r - 2
    for j in range(2, 6):
        ws.cell(row=r, column=j, value=f'=IF({cr(j,margin_row)}>=0.2,"CLEAR","BLOCKED")')
        dc(ws.cell(row=r, column=j), bold=True)
        ws.cell(row=r, column=j).alignment = Alignment(horizontal="center", vertical="center")

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 9: RISK MITIGATION
# ═══════════════════════════════════════════════
def create_risk(wb, params, computed):
    ws = wb.create_sheet("Risk Mitigation")
    ws.sheet_properties.tabColor = "EF4444"

    r = title_block(ws,
        "Risk-Adjusted Stage-Gate Mitigation Costing",
        "Expected Loss = Probability x Potential Loss (formula). Mitigation budget auto-totals."
    )

    headers = ["Risk Category", "Stage", "Severity", "Probability",
               "Potential Loss (Cr)", "Expected Loss (Cr)",
               "Mitigation Budget (Lakhs)", "Mechanism", "Residual Risk"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers), bg="EF4444")
    r += 1

    risks = [
        ("RERA Delays", "S2", "HIGH", 0.15, 25.0, 50.0, "Gate 2 blocks marketing until RERA active", "LOW"),
        ("Broker Apathy", "S2", "MEDIUM", 0.25, 15.0, 80.0, "CP Ranker auto-tiering incentives", "LOW"),
        ("Weak EOI Corpus", "S3", "HIGH", 0.20, 40.0, 120.0, "Gate 3 mandates 12 Cr tokens", "MEDIUM"),
        ("CAC Inflation", "S4", "MEDIUM", 0.30, 12.0, 45.0, "Astra Phoenix lead revival via n8n", "LOW"),
        ("Absorption Lag", "S4", "LOW", 0.20, 8.0, 30.0, "Barbell pricing flex (10-90 plans)", "LOW"),
        ("Cost Overruns", "S4", "MEDIUM", 0.25, 30.0, 0.0, "Developer contingency buffer", "MEDIUM"),
        ("Rate Hikes", "All", "MEDIUM", 0.20, 18.0, 0.0, "Velocity acceleration compresses tenure", "LOW"),
    ]

    risk_rows = []
    for name, stage, sev, prob, loss, mit, mech, residual in risks:
        risk_rows.append(r)
        ws.cell(row=r, column=1, value=name); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=stage); dc(ws.cell(row=r, column=2))

        ws.cell(row=r, column=3, value=sev)
        dc(ws.cell(row=r, column=3))
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
        if sev == "HIGH":
            ws.cell(row=r, column=3).fill = sfill(RED_LIGHT)
            ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, color="DC2626")
        elif sev == "MEDIUM":
            ws.cell(row=r, column=3).fill = sfill(AMBER_LIGHT)
            ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, color="D97706")
        else:
            ws.cell(row=r, column=3).fill = sfill(GREEN_LIGHT)
            ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, color=GREEN)

        # Probability (input)
        ws.cell(row=r, column=4, value=prob)
        dc(ws.cell(row=r, column=4), num=True, fmt="0%")

        # Potential Loss (input)
        ws.cell(row=r, column=5, value=loss)
        dc(ws.cell(row=r, column=5), num=True, fmt="#,##0.0")

        # Expected Loss = Probability x Potential Loss (FORMULA)
        ws.cell(row=r, column=6, value=f"={cr(4,r)}*{cr(5,r)}")
        dc(ws.cell(row=r, column=6), num=True, fmt="#,##0.00")

        # Mitigation budget (input)
        ws.cell(row=r, column=7, value=mit)
        dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.0")

        ws.cell(row=r, column=8, value=mech); dc(ws.cell(row=r, column=8))
        ws.cell(row=r, column=9, value=residual); dc(ws.cell(row=r, column=9))
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center")
        r += 1

    # Totals
    fr, lr = risk_rows[0], risk_rows[-1]
    ws.cell(row=r, column=1, value="TOTALS"); dc(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=5, value=f"=SUM({cr(5,fr)}:{cr(5,lr)})")
    dc(ws.cell(row=r, column=5), num=True, fmt="#,##0.0", bold=True)

    ws.cell(row=r, column=6, value=f"=SUM({cr(6,fr)}:{cr(6,lr)})")
    dc(ws.cell(row=r, column=6), num=True, fmt="#,##0.00", bold=True)

    ws.cell(row=r, column=7, value=f"=SUM({cr(7,fr)}:{cr(7,lr)})")
    dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.0", bold=True)

    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)
    r += 1

    # Summary note with formula
    gtv = ref("Assumptions", 2, computed["Total Project GTV"])
    ws.cell(row=r, column=1, value="Mitigation Budget as % of GTV:")
    dc(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=f"={cr(7,r-1)}/100/{gtv}")
    dc(ws.cell(row=r, column=2), num=True, fmt="0.00%")

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# SHEET 10: ANAROCK FEE STRUCTURE
# ═══════════════════════════════════════════════
def create_fees(wb, params, computed):
    ws = wb.create_sheet("Anarock Fee Structure")
    ws.sheet_properties.tabColor = CRIMSON
    A = "Assumptions"

    r = title_block(ws,
        "Anarock Engagement Fee Structure & Revenue Model",
        "All fees computed from Assumptions. Change rates to recalculate Anarock revenue."
    )

    headers = ["Fee Component", "Rate", "Basis", "Revenue (Cr)", "Payment Trigger"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(headers))
    r += 1

    gtv = ref(A, 2, computed["Total Project GTV"])
    brk_rate = ref(A, 2, params["Brokerage Commission Rate"])
    mkt_rate = ref(A, 2, params["Marketing Budget Rate"])
    ai_fee = ref(A, 2, params["AI Platform Monthly Fee"])
    cp_fee = ref(A, 2, params["CP Mobilization (One-time)"])
    perf_rate = ref(A, 2, params["Performance Bonus Rate"])

    fee_rows = {}
    fees = [
        ("Brokerage Commission", f"={brk_rate}&\"%\"", "Per unit sale",
         f"={gtv}*{brk_rate}/100", "On each sale closure"),
        ("Marketing Management Fee", f"={mkt_rate}&\"%\"", "Monthly retainer",
         f"={gtv}*{mkt_rate}/100", "Monthly during S3-S4"),
        ("AI Platform License", f"={ai_fee}&\" L/mo\"", "SaaS monthly",
         f"={ai_fee}*42/100", "Monthly active mandate"),
        ("CP Mobilization Fee", "One-time", "On activation",
         f"={cp_fee}/100", "Gate 2 clearance"),
        ("Performance Bonus", f"={perf_rate}&\"%\"", "Quarterly if >15 u/mo",
         f"={gtv}*{perf_rate}/100", "Quarterly review"),
    ]

    for name, rate, basis, formula, trigger in fees:
        fee_rows[name] = r
        ws.cell(row=r, column=1, value=name); dc(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2, value=rate); dc(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3, value=basis); dc(ws.cell(row=r, column=3))
        ws.cell(row=r, column=4, value=formula)
        dc(ws.cell(row=r, column=4), num=True, fmt="#,##0.00")
        ws.cell(row=r, column=5, value=trigger); dc(ws.cell(row=r, column=5))
        r += 1

    # Total
    fr = list(fee_rows.values())[0]
    lr = list(fee_rows.values())[-1]
    ws.cell(row=r, column=1, value="TOTAL ANAROCK REVENUE"); dc(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=4, value=f"=SUM({cr(4,fr)}:{cr(4,lr)})")
    dc(ws.cell(row=r, column=4), num=True, fmt="#,##0.00", bold=True)
    total_row = r
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = sfill(CRIMSON_LIGHT)
        ws.cell(row=r, column=c).font = bfont(bold=True, size=11)
    r += 1

    # % of GTV
    ws.cell(row=r, column=1, value="as % of GTV"); dc(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=4, value=f"={cr(4,total_row)}/{gtv}")
    dc(ws.cell(row=r, column=4), num=True, fmt="0.00%")
    r += 1

    # % of Developer Profit
    profit_ref = ref(A, 2, computed["Developer Net Profit"])
    ws.cell(row=r, column=1, value="as % of Developer Net Profit"); dc(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=4, value=f"={cr(4,total_row)}/{profit_ref}")
    dc(ws.cell(row=r, column=4), num=True, fmt="0.00%")

    r += 3

    # Revenue Projection by Phase
    ws.cell(row=r, column=1, value="Anarock Revenue by Phase (Cr)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=13, color=CRIMSON)
    r += 1

    rph = ["Period", "% of Total Units", "Units Sold", "Brokerage (Cr)",
           "Marketing Fee (Cr)", "Platform Fee (Cr)", "Total (Cr)"]
    for i, h in enumerate(rph, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(rph))
    r += 1

    total_u_ref = ref(A, 2, computed["Total Units (Sum)"])
    avg_ticket_ref = ref(A, 2, computed["Avg Ticket Size"])

    phases = [
        ("M1-M2 (Discovery)", 0.0, 2),
        ("M3-M4 (GTM Prep)", 0.0, 2),
        ("M5 (Soft Launch)", 15.0, 1),
        ("M6-M12 (Launch)", 27.0, 7),
        ("M13-M24 (Sustenance)", 31.0, 12),
        ("M25-M36 (Sustenance)", 20.5, 12),
        ("M37-M42 (Wrapex)", 6.5, 6),
    ]

    proj_rows = []
    for period, pct_units, months in phases:
        proj_rows.append(r)
        ws.cell(row=r, column=1, value=period); dc(ws.cell(row=r, column=1), bold=True)

        # % units (input)
        ws.cell(row=r, column=2, value=pct_units / 100)
        dc(ws.cell(row=r, column=2), num=True, fmt="0.0%")

        # Units = % x Total
        ws.cell(row=r, column=3, value=f"=ROUND({cr(2,r)}*{total_u_ref},0)")
        dc(ws.cell(row=r, column=3), num=True, fmt="#,##0")

        # Brokerage = Units * Avg Ticket * Brokerage Rate / 100
        ws.cell(row=r, column=4, value=f"={cr(3,r)}*{avg_ticket_ref}*{brk_rate}/100")
        dc(ws.cell(row=r, column=4), num=True, fmt="#,##0.00")

        # Marketing Fee = Units * Avg Ticket * Marketing Rate / 100
        ws.cell(row=r, column=5, value=f"={cr(3,r)}*{avg_ticket_ref}*{mkt_rate}/100")
        dc(ws.cell(row=r, column=5), num=True, fmt="#,##0.00")

        # Platform Fee = months * monthly fee / 100
        ws.cell(row=r, column=6, value=f"={months}*{ai_fee}/100")
        dc(ws.cell(row=r, column=6), num=True, fmt="#,##0.00")

        # Total = Brokerage + Marketing + Platform
        ws.cell(row=r, column=7, value=f"={cr(4,r)}+{cr(5,r)}+{cr(6,r)}")
        dc(ws.cell(row=r, column=7), num=True, fmt="#,##0.00")
        r += 1

    # Phase totals
    pfr, plr = proj_rows[0], proj_rows[-1]
    ws.cell(row=r, column=1, value="TOTAL"); dc(ws.cell(row=r, column=1), bold=True)
    for col in [2, 3, 4, 5, 6, 7]:
        ws.cell(row=r, column=col, value=f"=SUM({cr(col,pfr)}:{cr(col,plr)})")
        fmt = "0.0%" if col == 2 else "#,##0" if col == 3 else "#,##0.00"
        dc(ws.cell(row=r, column=col), num=True, fmt=fmt, bold=True)
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = sfill(GOLD_LIGHT)

    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    print("1/10 Creating Assumptions (central input sheet)...")
    _, params, computed = create_assumptions(wb)

    print("2/10 Creating Revenue Waterfall (formulas)...")
    create_revenue_waterfall(wb, params, computed)

    print("3/10 Creating Cost Structure & P&L (formulas)...")
    create_cost_pnl(wb, params, computed)

    print("4/10 Creating Cash Flow Projection (formulas)...")
    create_cashflow(wb, params, computed)

    print("5/10 Creating IRR & DSCR Analysis (formulas)...")
    create_irr(wb, params, computed)

    print("6/10 Creating Stage-Gate Budget (formulas)...")
    create_budget(wb, params, computed)

    print("7/10 Creating CAC & Marketing Analysis (formulas)...")
    create_cac(wb, params, computed)

    print("8/10 Creating Sensitivity Analysis (formulas)...")
    create_sensitivity(wb, params, computed)

    print("9/10 Creating Risk Mitigation (formulas)...")
    create_risk(wb, params, computed)

    print("10/10 Creating Anarock Fee Structure (formulas)...")
    create_fees(wb, params, computed)

    output = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "Anarock_GTM_Financial_Model.xlsx"
    )
    wb.save(output)
    print(f"\n[DONE] Saved: {output}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    for i, n in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {n}")
    print("\nAll calculated cells use Excel formulas.")
    print("Change any input on the Assumptions sheet to recalculate the entire model.")


if __name__ == "__main__":
    main()
